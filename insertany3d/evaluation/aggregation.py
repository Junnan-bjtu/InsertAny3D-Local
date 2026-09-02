"""Task, scene, and batch aggregation for cached GPTEval responses."""

from __future__ import annotations

import csv
import json
import os
import statistics
import threading
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable, Mapping

from .gpteval import (
    GPTEvalRequest,
    ResponseCache,
    evaluation_config_sha256,
    normalize_dimensions,
    validate_gpteval_request,
)
from .manifests import EvaluationError, EvaluationManifest, validate_manifest_collection


def aggregate_gpteval(
    manifests: Iterable[EvaluationManifest],
    requests: Iterable[GPTEvalRequest],
    cache: ResponseCache,
    *,
    expected_tasks_per_method: int = 60,
    expected_scenes_per_method: int = 12,
    expected_tasks_per_scene: int = 5,
) -> dict[str, Any]:
    """Aggregate available cache records while preserving formal denominators."""

    records = list(manifests)
    request_list = list(requests)
    collection = validate_manifest_collection(records)
    for name, value in (
        ("expected_tasks_per_method", expected_tasks_per_method),
        ("expected_scenes_per_method", expected_scenes_per_method),
        ("expected_tasks_per_scene", expected_tasks_per_scene),
    ):
        if not isinstance(value, int) or isinstance(value, bool) or value < 1:
            raise EvaluationError(f"{name} 必须是正整数")
    if not request_list:
        raise EvaluationError("GPTEval 请求计划为空")

    manifest_by_key = {record.task_key: record for record in records}
    requests_by_task: dict[str, list[GPTEvalRequest]] = defaultdict(list)
    settings: set[tuple[str, str, str, tuple[str, ...]]] = set()
    request_keys: set[str] = set()
    for request in request_list:
        if request.request_key in request_keys:
            raise EvaluationError(f"GPTEval 请求计划包含重复 requestKey: {request.request_key}")
        request_keys.add(request.request_key)
        manifest = manifest_by_key.get(request.task_key)
        if manifest is None:
            raise EvaluationError(f"GPTEval 请求没有对应评测清单: {request.task_key}")
        validate_gpteval_request(request, manifest)
        requests_by_task[request.task_key].append(request)
        settings.add(
            (
                request.evaluator_version,
                request.model,
                request.rubric_sha256,
                request.dimensions,
            )
        )
    if set(requests_by_task) != set(manifest_by_key):
        missing = sorted(set(manifest_by_key) - set(requests_by_task))
        raise EvaluationError(f"部分评测清单没有请求计划: {missing[:3]}")
    if len(settings) != 1:
        raise EvaluationError("一次汇总不能混用 evaluator 版本、模型或评分规则")

    repeat_sets = {
        tuple(sorted(request.repeat_index for request in task_requests))
        for task_requests in requests_by_task.values()
    }
    if len(repeat_sets) != 1:
        raise EvaluationError("所有任务必须使用相同的 GPTEval repeat 计划")
    repeats = next(iter(repeat_sets))
    if repeats != tuple(range(len(repeats))):
        raise EvaluationError("repeatIndex 必须从 0 开始连续编号")

    evaluator_version, model, rubric_sha, dimensions = next(iter(settings))
    task_scores_by_method: dict[str, list[dict[str, Any]]] = defaultdict(list)
    ready_responses = 0
    missing_request_keys: list[str] = []
    for manifest in sorted(
        records, key=lambda item: (item.method_id, item.project_id, item.task_id)
    ):
        task_requests = sorted(
            requests_by_task[manifest.task_key], key=lambda item: item.repeat_index
        )
        cached = []
        for request in task_requests:
            response = cache.get(request)
            if response is None:
                missing_request_keys.append(request.request_key)
            else:
                cached.append(response)
                ready_responses += 1
        task_scores_by_method[manifest.method_id].append(
            _task_score(
                manifest,
                cached,
                expected_repeats=len(repeats),
                dimensions=dimensions,
            )
        )

    methods: dict[str, Any] = {}
    all_task_scores: list[dict[str, Any]] = []
    all_scene_scores: list[dict[str, Any]] = []
    for method_id in sorted(task_scores_by_method):
        task_scores = task_scores_by_method[method_id]
        scene_scores = _scene_scores(
            task_scores,
            expected_tasks_per_scene,
            dimensions=dimensions,
        )
        ready_tasks = sum(item["status"] == "ready" for item in task_scores)
        ready_scenes = sum(item["status"] == "ready" for item in scene_scores)
        method_ready = (
            len(task_scores) == expected_tasks_per_method
            and ready_tasks == expected_tasks_per_method
            and len(scene_scores) == expected_scenes_per_method
            and ready_scenes == expected_scenes_per_method
        )
        methods[method_id] = {
            "status": "ready" if method_ready else "partial",
            "completion": {
                "expectedTasks": expected_tasks_per_method,
                "discoveredTasks": len(task_scores),
                "readyTasks": ready_tasks,
                "expectedScenes": expected_scenes_per_method,
                "discoveredScenes": len(scene_scores),
                "readyScenes": ready_scenes,
                "missingTaskSlots": max(0, expected_tasks_per_method - len(task_scores)),
                "missingSceneSlots": max(0, expected_scenes_per_method - len(scene_scores)),
            },
            "taskScores": task_scores,
            "sceneScores": scene_scores,
            "batchScore": _batch_score(
                task_scores,
                scene_scores,
                expected_tasks=expected_tasks_per_method,
                expected_scenes=expected_scenes_per_method,
                dimensions=dimensions,
            ),
        }
        all_task_scores.extend(task_scores)
        all_scene_scores.extend(scene_scores)

    expected_responses = expected_tasks_per_method * len(methods) * len(repeats)
    status = "ready" if methods and all(
        item["status"] == "ready" for item in methods.values()
    ) else "partial"
    return {
        "schemaVersion": 1,
        "kind": "insertany3d.gpteval-summary",
        "status": status,
        "batchId": collection["batchId"],
        "metric": "gpteval",
        "evaluatorVersion": evaluator_version,
        "model": model,
        "rubricSha256": rubric_sha,
        "dimensions": list(dimensions),
        "comparisonConfigSha256": evaluation_config_sha256(
            collection["comparisonConfigSha256"], dimensions
        ),
        "repeats": len(repeats),
        "completion": {
            "methods": len(methods),
            "expectedTaskResults": expected_tasks_per_method * len(methods),
            "discoveredTaskResults": len(all_task_scores),
            "readyTaskResults": sum(item["status"] == "ready" for item in all_task_scores),
            "expectedSceneResults": expected_scenes_per_method * len(methods),
            "discoveredSceneResults": len(all_scene_scores),
            "readySceneResults": sum(item["status"] == "ready" for item in all_scene_scores),
            "expectedResponses": expected_responses,
            "discoveredResponses": len(request_list),
            "readyResponses": ready_responses,
            "missingResponses": max(0, expected_responses - ready_responses),
            "missingPlannedResponses": len(missing_request_keys),
        },
        "missingRequestKeys": sorted(missing_request_keys),
        "methods": methods,
    }


def write_gpteval_summary(
    output_dir: str | Path,
    summary: Mapping[str, Any],
) -> dict[str, Path]:
    """Write machine-readable outputs and a human-readable XLSX workbook."""

    root = Path(output_dir)
    root.mkdir(parents=True, exist_ok=True)
    _atomic_write_json(root / "batch_summary.json", summary)

    task_rows = []
    scene_rows = []
    for method in summary.get("methods", {}).values():
        task_rows.extend(method.get("taskScores", []))
        scene_rows.extend(method.get("sceneScores", []))
    _atomic_write_jsonl(root / "task_scores.jsonl", task_rows)
    dimensions = normalize_dimensions(summary.get("dimensions"))
    _write_scene_csv(root / "scene_scores.csv", scene_rows, dimensions=dimensions)
    _write_summary_xlsx(
        root / "gpteval_summary.xlsx",
        summary,
        task_rows,
        scene_rows,
        dimensions=dimensions,
    )
    return {
        "batch": root / "batch_summary.json",
        "tasks": root / "task_scores.jsonl",
        "scenes": root / "scene_scores.csv",
        "xlsx": root / "gpteval_summary.xlsx",
    }


def _task_score(
    manifest: EvaluationManifest,
    responses: list[Mapping[str, Any]],
    *,
    expected_repeats: int,
    dimensions: tuple[str, ...],
) -> dict[str, Any]:
    scores = {}
    for dimension in dimensions:
        stats = _score_stats(
            [float(response["scores"][dimension]["score"]) for response in responses],
            expected_count=expected_repeats,
        )
        stats["reasons"] = [
            str(response["scores"][dimension]["reason"]) for response in responses
        ]
        scores[dimension] = stats
    input_images = [
        {
            "viewId": str(view["viewId"]),
            "original": str(view["original"]["path"]),
            "inserted": str(view["inserted"]["path"]),
        }
        for view in manifest.data["views"]
    ]
    return {
        "projectId": manifest.project_id,
        "scenePath": manifest.scene_path,
        "taskId": manifest.task_id,
        "runId": manifest.run_id,
        "methodId": manifest.method_id,
        "taskPrompt": manifest.data["taskPrompt"],
        "inputImages": input_images,
        "manifestSha256": manifest.manifest_sha256,
        "status": "ready" if len(responses) == expected_repeats else "partial",
        "completion": {
            "expectedRepeats": expected_repeats,
            "readyRepeats": len(responses),
            "missingRepeats": expected_repeats - len(responses),
        },
        "scores": scores,
        "totalScore": _dimension_average(scores, dimensions),
    }


def _scene_scores(
    task_scores: list[dict[str, Any]],
    expected_tasks_per_scene: int,
    *,
    dimensions: tuple[str, ...],
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for task in task_scores:
        grouped[(task["projectId"], task["scenePath"], task["methodId"])].append(task)

    result = []
    for (project_id, scene_path, method_id), tasks in sorted(grouped.items()):
        ready_tasks = [task for task in tasks if task["status"] == "ready"]
        scored_tasks = [
            task
            for task in tasks
            if all(task["scores"][dimension]["count"] > 0 for dimension in dimensions)
        ]
        scores = {
            dimension: _score_stats(
                [task["scores"][dimension]["mean"] for task in scored_tasks],
                expected_count=expected_tasks_per_scene,
            )
            for dimension in dimensions
        }
        result.append(
            {
                "projectId": project_id,
                "scenePath": scene_path,
                "methodId": method_id,
                "status": (
                    "ready"
                    if len(tasks) == expected_tasks_per_scene
                    and len(ready_tasks) == expected_tasks_per_scene
                    else "partial"
                ),
                "completion": {
                    "expectedTasks": expected_tasks_per_scene,
                    "discoveredTasks": len(tasks),
                    "readyTasks": len(ready_tasks),
                    "scoredTasks": len(scored_tasks),
                    "missingTaskSlots": max(0, expected_tasks_per_scene - len(tasks)),
                },
                "scores": scores,
                "totalScore": _dimension_average(scores, dimensions),
            }
        )
    return result


def _batch_score(
    task_scores: list[dict[str, Any]],
    scene_scores: list[dict[str, Any]],
    *,
    expected_tasks: int,
    expected_scenes: int,
    dimensions: tuple[str, ...],
) -> dict[str, Any]:
    scored_tasks = [
        task
        for task in task_scores
        if all(task["scores"][dimension]["count"] > 0 for dimension in dimensions)
    ]
    scored_scenes = [
        scene
        for scene in scene_scores
        if all(scene["scores"][dimension]["count"] > 0 for dimension in dimensions)
    ]
    scene_macro = {
        dimension: _score_stats(
            [scene["scores"][dimension]["mean"] for scene in scored_scenes],
            expected_count=expected_scenes,
        )
        for dimension in dimensions
    }
    task_macro = {
        dimension: _score_stats(
            [task["scores"][dimension]["mean"] for task in scored_tasks],
            expected_count=expected_tasks,
        )
        for dimension in dimensions
    }
    return {
        "sceneMacroAverage": scene_macro,
        "taskMacroAverage": task_macro,
        "totalScore": _dimension_average(scene_macro, dimensions),
    }


def _dimension_average(
    scores: Mapping[str, Mapping[str, Any]],
    dimensions: tuple[str, ...],
) -> float | None:
    values = [scores[dimension].get("mean") for dimension in dimensions]
    if any(value is None for value in values):
        return None
    numeric_values = [float(value) for value in values if value is not None]
    return round(statistics.fmean(numeric_values), 6)


def _score_stats(values: list[float], *, expected_count: int) -> dict[str, Any]:
    if not values:
        return {
            "mean": None,
            "standardDeviation": None,
            "count": 0,
            "expectedCount": expected_count,
            "missingCount": expected_count,
        }
    return {
        "mean": round(statistics.fmean(values), 6),
        "standardDeviation": (
            round(statistics.stdev(values), 6) if len(values) > 1 else 0.0
        ),
        "count": len(values),
        "expectedCount": expected_count,
        "missingCount": max(0, expected_count - len(values)),
    }


def _atomic_write_json(path: Path, value: Mapping[str, Any]) -> None:
    _atomic_write_text(
        path,
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )


def _atomic_write_jsonl(path: Path, rows: list[Mapping[str, Any]]) -> None:
    text = "".join(
        json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in rows
    )
    _atomic_write_text(path, text)


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(
        f"{path.name}.{os.getpid()}.{threading.get_ident()}.tmp"
    )
    try:
        with temporary.open("w", encoding="utf-8", newline="") as stream:
            stream.write(text)
            stream.flush()
            os.fsync(stream.fileno())
        temporary.replace(path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _write_scene_csv(
    path: Path,
    rows: list[Mapping[str, Any]],
    *,
    dimensions: tuple[str, ...],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(
        f"{path.name}.{os.getpid()}.{threading.get_ident()}.tmp"
    )
    try:
        with temporary.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream)
            writer.writerow(
                [
                    "project_id",
                    "scene_path",
                    "method_id",
                    "status",
                    "ready_tasks",
                    "expected_tasks",
                    *dimensions,
                    "total_score",
                ]
            )
            for row in rows:
                writer.writerow(
                    [
                        row["projectId"],
                        row["scenePath"],
                        row["methodId"],
                        row["status"],
                        row["completion"]["readyTasks"],
                        row["completion"]["expectedTasks"],
                        *[row["scores"][dimension]["mean"] for dimension in dimensions],
                        row.get("totalScore"),
                    ]
                )
            stream.flush()
            os.fsync(stream.fileno())
        temporary.replace(path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _write_summary_xlsx(
    path: Path,
    summary: Mapping[str, Any],
    task_rows: list[Mapping[str, Any]],
    scene_rows: list[Mapping[str, Any]],
    *,
    dimensions: tuple[str, ...],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise EvaluationError("生成 XLSX 需要 openpyxl；请重新安装项目依赖") from exc

    labels = {
        "visual_quality": "视觉质量",
        "insertion_rationality": "插入合理性",
        "geometric_accuracy": "几何准确性",
    }
    workbook = Workbook()
    scenes_sheet = workbook.active
    if scenes_sheet is None:
        raise EvaluationError("openpyxl 未创建默认工作表")
    scenes_sheet.title = "场景汇总"
    scene_header = [
        "项目",
        "场景",
        "方法",
        "状态",
        "已完成任务",
        "预计任务",
        *[labels[dimension] for dimension in dimensions],
        "场景总分",
    ]
    _append_xlsx_row(scenes_sheet, scene_header)
    for row in scene_rows:
        _append_xlsx_row(
            scenes_sheet,
            [
                row["projectId"],
                row["scenePath"],
                row["methodId"],
                row["status"],
                row["completion"]["readyTasks"],
                row["completion"]["expectedTasks"],
                *[row["scores"][dimension]["mean"] for dimension in dimensions],
                row.get("totalScore"),
            ]
        )
    for method_id, method in sorted(summary.get("methods", {}).items()):
        batch_score = method["batchScore"]
        _append_xlsx_row(
            scenes_sheet,
            [
                "总体",
                "总体",
                method_id,
                method["status"],
                method["completion"]["readyTasks"],
                method["completion"]["expectedTasks"],
                *[
                    batch_score["sceneMacroAverage"][dimension]["mean"]
                    for dimension in dimensions
                ],
                batch_score.get("totalScore"),
            ]
        )

    tasks_sheet = workbook.create_sheet("任务明细")
    image_columns = [
        f"{kind}_{view_id}"
        for kind in ("原图", "插入图")
        for view_id in (
            "low_left",
            "low_center",
            "low_right",
            "high_left",
            "high_center",
            "high_right",
        )
    ]
    # Keep the two requested quality dimensions next to the task identity;
    # optional dimensions and operational/image columns follow afterwards.
    primary_dimensions = tuple(
        dimension
        for dimension in ("visual_quality", "geometric_accuracy")
        if dimension in dimensions
    )
    remaining_dimensions = tuple(
        dimension for dimension in dimensions if dimension not in primary_dimensions
    )
    task_header = ["项目", "任务", "插入要求"]
    for dimension in primary_dimensions:
        task_header.extend((f"{labels[dimension]}分数", f"{labels[dimension]}理由"))
    task_header.extend(("场景", "运行ID", "方法"))
    for dimension in remaining_dimensions:
        task_header.extend((f"{labels[dimension]}分数", f"{labels[dimension]}理由"))
    task_header.extend((*image_columns, "任务总分", "状态"))
    _append_xlsx_row(tasks_sheet, task_header)
    for row in task_rows:
        images = {
            (kind, str(item["viewId"])): str(item[kind])
            for item in row.get("inputImages", [])
            for kind in ("original", "inserted")
        }
        values: list[Any] = [row["projectId"], row["taskId"], row["taskPrompt"]]
        for dimension in primary_dimensions:
            score = row["scores"][dimension]
            values.extend((score["mean"], " | ".join(score.get("reasons", []))))
        values.extend((row["scenePath"], row["runId"], row["methodId"]))
        for dimension in remaining_dimensions:
            score = row["scores"][dimension]
            values.extend((score["mean"], " | ".join(score.get("reasons", []))))
        for kind in ("original", "inserted"):
            for view_id in (
                "low_left",
                "low_center",
                "low_right",
                "high_left",
                "high_center",
                "high_right",
            ):
                values.append(images.get((kind, view_id)))
        values.extend((row.get("totalScore"), row["status"]))
        _append_xlsx_row(tasks_sheet, values)

    for sheet in (scenes_sheet, tasks_sheet):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column_index, column_cells in enumerate(sheet.columns, start=1):
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(width + 2, 10), 48
            )

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(
        f".{path.stem}.{os.getpid()}.{threading.get_ident()}.tmp.xlsx"
    )
    try:
        workbook.save(temporary)
        temporary.replace(path)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()


def _append_xlsx_row(sheet: Any, values: Iterable[Any]) -> None:
    """Append data as values, never as formulas supplied by task metadata."""

    sheet.append(list(values))
    for cell in sheet[sheet.max_row]:
        if isinstance(cell.value, str):
            cell.data_type = "s"
