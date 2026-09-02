from __future__ import annotations

import hashlib
import json
import multiprocessing
import queue
import struct
import tempfile
import threading
import time
import unittest
import zlib
from concurrent.futures import ThreadPoolExecutor
from dataclasses import replace
from pathlib import Path

from insertany3d.contracts.models import canonical_sha256
from insertany3d.evaluation import (
    DEFAULT_DIMENSIONS,
    DIMENSIONS,
    SUPPORTED_DIMENSIONS,
    EvaluationError,
    ResponseCache,
    adapt_gpteval_response,
    aggregate_gpteval,
    build_request_body,
    discover_evaluation_manifests,
    execute_gpteval_requests,
    fixed_fake_response,
    load_evaluation_manifest,
    make_request_key,
    normalize_dimensions,
    pending_gpteval_requests,
    plan_gpteval_requests,
    require_supported_evaluator,
    rubric_sha256,
    write_gpteval_summary,
)


VIEW_LAYOUT = (
    ("low_left", 10, -1),
    ("low_center", 10, 0),
    ("low_right", 10, 1),
    ("high_left", 40, -1),
    ("high_center", 40, 0),
    ("high_right", 40, 1),
)


class EvaluationTests(unittest.TestCase):
    def test_unity_camera_pose_and_matrix_fields_are_compatible(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_unity_camera_") as directory:
            manifest_path = _write_manifest(
                Path(directory), project_index=1, task_index=1
            )
            camera_path = manifest_path.parent / "cameras" / "low_left.camera.json"
            camera = json.loads(camera_path.read_text(encoding="utf-8"))
            self.assertIn("cameraToWorld", camera)
            self.assertEqual(len(camera["cameraToWorldMatrix"]), 16)
            self.assertEqual(len(camera["projectionMatrix"]), 16)

            record = load_evaluation_manifest(manifest_path)

            self.assertEqual(record.data["views"][0]["viewId"], "low_left")

    def test_60_fake_manifests_resume_and_three_level_aggregation(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_") as directory:
            root = Path(directory)
            _create_batch(root, project_count=12, task_count=5)
            manifests = discover_evaluation_manifests(root)
            self.assertEqual(len(manifests), 60)
            self.assertEqual({item.data["viewConfig"]["yawOffsetDegrees"] for item in manifests}, {24})

            rubric_hash = rubric_sha256("fixed rubric v1")
            requests = plan_gpteval_requests(
                manifests,
                evaluator_version="gpteval-library-v1",
                model="offline-fake-model",
                rubric_sha256_value=rubric_hash,
            )
            self.assertEqual(len(requests), 60)
            cache = ResponseCache(root / "cache")
            called: list[str] = []

            def evaluator(request):
                called.append(request.request_key)
                return fixed_fake_response(8)

            interrupted = execute_gpteval_requests(
                requests, cache, evaluator, limit=17
            )
            self.assertEqual(
                interrupted,
                {"planned": 17, "completed": 17, "skipped": 0, "failed": 0},
            )
            self.assertEqual(len(pending_gpteval_requests(requests, cache)), 43)

            resumed = execute_gpteval_requests(requests, cache, evaluator)
            self.assertEqual(resumed["planned"], 43)
            self.assertEqual(resumed["completed"], 43)
            self.assertEqual(resumed["skipped"], 17)
            self.assertEqual(resumed["failed"], 0)
            self.assertEqual(len(called), 60)

            summary = aggregate_gpteval(manifests, requests, cache)
            self.assertEqual(summary["status"], "ready")
            self.assertEqual(summary["completion"]["readyTaskResults"], 60)
            self.assertEqual(summary["completion"]["readySceneResults"], 12)
            method = summary["methods"]["insertany3d-main"]
            self.assertEqual(len(method["taskScores"]), 60)
            self.assertEqual(len(method["sceneScores"]), 12)
            for dimension in DIMENSIONS:
                self.assertEqual(
                    method["batchScore"]["sceneMacroAverage"][dimension]["mean"],
                    8.0,
                )

            output_dir = root / "summary"
            write_gpteval_summary(output_dir, summary)
            self.assertTrue((output_dir / "batch_summary.json").is_file())
            self.assertEqual(
                len((output_dir / "task_scores.jsonl").read_text(encoding="utf-8").splitlines()),
                60,
            )
            from openpyxl import load_workbook

            workbook = load_workbook(output_dir / "gpteval_summary.xlsx", read_only=True)
            self.assertEqual(workbook.sheetnames, ["场景汇总", "任务明细"])
            self.assertEqual(workbook["场景汇总"].max_row, 14)
            self.assertEqual(workbook["任务明细"].max_row, 61)
            scene_headers = [cell.value for cell in workbook["场景汇总"][1]]
            task_headers = [cell.value for cell in workbook["任务明细"][1]]
            self.assertEqual(
                task_headers[:7],
                [
                    "项目",
                    "任务",
                    "插入要求",
                    "视觉质量分数",
                    "视觉质量理由",
                    "几何准确性分数",
                    "几何准确性理由",
                ],
            )
            self.assertIn("视觉质量", scene_headers)
            self.assertIn("几何准确性", scene_headers)
            self.assertNotIn("插入合理性", scene_headers)
            self.assertEqual(
                len([header for header in task_headers if str(header).startswith("原图_")]),
                6,
            )
            self.assertEqual(
                len([header for header in task_headers if str(header).startswith("插入图_")]),
                6,
            )
            first_task = [cell.value for cell in workbook["任务明细"][2]]
            image_indexes = [
                index
                for index, header in enumerate(task_headers)
                if str(header).startswith(("原图_", "插入图_"))
            ]
            self.assertEqual(len(image_indexes), 12)
            for index in image_indexes:
                filename = str(first_task[index])
                self.assertFalse(Path(filename).is_absolute())
                self.assertNotRegex(filename, r"^[A-Za-z]:[\\/]")
            workbook.close()

            cache.response_path(requests[0]).unlink()
            partial = aggregate_gpteval(manifests, requests, cache)
            self.assertEqual(partial["status"], "partial")
            self.assertEqual(partial["completion"]["expectedResponses"], 60)
            self.assertEqual(partial["completion"]["readyResponses"], 59)
            self.assertEqual(partial["completion"]["missingResponses"], 1)
            self.assertEqual(partial["completion"]["readyTaskResults"], 59)
            self.assertEqual(partial["completion"]["readySceneResults"], 11)
            self.assertEqual(len(partial["missingRequestKeys"]), 1)

            recovery_calls = []
            recovered = execute_gpteval_requests(
                requests,
                cache,
                lambda request: recovery_calls.append(request.request_key)
                or fixed_fake_response(8),
            )
            self.assertEqual(
                recovered,
                {"planned": 1, "completed": 1, "skipped": 59, "failed": 0},
            )
            self.assertEqual(recovery_calls, [requests[0].request_key])

    def test_xlsx_writes_untrusted_strings_as_plain_text(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_xlsx_text_") as directory:
            root = Path(directory)
            manifest = load_evaluation_manifest(
                _write_manifest(root / "task", project_index=1, task_index=1)
            )
            request = plan_gpteval_requests(
                [manifest],
                evaluator_version="v1",
                model="offline",
                rubric_sha256_value=rubric_sha256("rubric"),
            )[0]
            cache = ResponseCache(root / "cache")
            cache.store(
                request,
                {
                    dimension: {"score": 7, "reason": '=HYPERLINK("https://invalid")'}
                    for dimension in DEFAULT_DIMENSIONS
                },
            )
            summary = aggregate_gpteval(
                [manifest],
                [request],
                cache,
                expected_tasks_per_method=1,
                expected_scenes_per_method=1,
                expected_tasks_per_scene=1,
            )
            task = summary["methods"]["insertany3d-main"]["taskScores"][0]
            task["taskPrompt"] = "+not-a-formula"
            task["inputImages"][0]["original"] = "@not-a-formula.png"
            scene = summary["methods"]["insertany3d-main"]["sceneScores"][0]
            scene["scenePath"] = "-not-a-formula"

            output = root / "summary"
            write_gpteval_summary(output, summary)

            from openpyxl import load_workbook

            workbook = load_workbook(output / "gpteval_summary.xlsx")
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            self.assertEqual(cell.data_type, "s", cell.coordinate)
            workbook.close()

    def test_preflight_rejects_file_camera_and_collection_mismatches(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_bad_") as directory:
            root = Path(directory)
            manifest_path = _write_manifest(root / "valid", project_index=1, task_index=1)
            record = load_evaluation_manifest(manifest_path)
            self.assertEqual(record.data["viewConfig"]["yawOffsetDegrees"], 24)

            first_image = manifest_path.parent / "original" / "low_left.png"
            first_image.write_bytes(_png(1, 1, (255, 0, 0)))
            with self.assertRaisesRegex(EvaluationError, "SHA-256"):
                load_evaluation_manifest(manifest_path)

            dimension_path = _write_manifest(
                root / "bad_dimension", project_index=2, task_index=1
            )
            dimension_manifest = json.loads(dimension_path.read_text(encoding="utf-8"))
            image_path = dimension_path.parent / "inserted" / "high_right.png"
            image_path.write_bytes(_png(2, 1, (0, 255, 0)))
            dimension_manifest["views"][-1]["inserted"]["sha256"] = _sha(image_path)
            _write_json(dimension_path, dimension_manifest)
            with self.assertRaisesRegex(EvaluationError, "尺寸"):
                load_evaluation_manifest(dimension_path)

            camera_path = _write_manifest(
                root / "bad_camera", project_index=3, task_index=1
            )
            camera_manifest = json.loads(camera_path.read_text(encoding="utf-8"))
            camera_file = camera_path.parent / "cameras" / "low_left.camera.json"
            camera = json.loads(camera_file.read_text(encoding="utf-8"))
            camera["viewId"] = "low_right"
            _write_json(camera_file, camera)
            camera_manifest["views"][0]["camera"]["sha256"] = _sha(camera_file)
            _write_json(camera_path, camera_manifest)
            with self.assertRaisesRegex(EvaluationError, "viewId"):
                load_evaluation_manifest(camera_path)

            mixed = root / "mixed"
            _write_manifest(mixed / "one", project_index=4, task_index=1, yaw=24)
            _write_manifest(mixed / "two", project_index=5, task_index=1, yaw=12)
            with self.assertRaisesRegex(EvaluationError, "viewConfig"):
                discover_evaluation_manifests(mixed)

    def test_preflight_rejects_symlinks_and_invalid_camera_payloads(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_security_") as directory:
            root = Path(directory)
            manifest_path = _write_manifest(root / "missing_pose", project_index=1, task_index=1)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            camera_path = manifest_path.parent / "cameras" / "low_left.camera.json"
            camera = json.loads(camera_path.read_text(encoding="utf-8"))
            del camera["cameraToWorldMatrix"]
            _write_json(camera_path, camera)
            manifest["views"][0]["camera"]["sha256"] = _sha(camera_path)
            _write_json(manifest_path, manifest)
            evaluator_calls = []
            with self.assertRaisesRegex(EvaluationError, "cameraToWorldMatrix"):
                records = [load_evaluation_manifest(manifest_path)]
                requests = plan_gpteval_requests(
                    records,
                    evaluator_version="v1",
                    model="offline",
                    rubric_sha256_value=rubric_sha256("rubric"),
                )
                execute_gpteval_requests(
                    requests,
                    ResponseCache(root / "unused-cache"),
                    lambda request: evaluator_calls.append(request.request_key),
                )
            self.assertEqual(evaluator_calls, [])

            short_matrix_path = _write_manifest(
                root / "short_matrix", project_index=1, task_index=2
            )
            short_matrix_manifest = json.loads(short_matrix_path.read_text(encoding="utf-8"))
            short_matrix_camera_path = (
                short_matrix_path.parent / "cameras" / "low_left.camera.json"
            )
            short_matrix_camera = json.loads(
                short_matrix_camera_path.read_text(encoding="utf-8")
            )
            short_matrix_camera["projectionMatrix"] = [1, 0, 0, 1]
            _write_json(short_matrix_camera_path, short_matrix_camera)
            short_matrix_manifest["views"][0]["camera"]["sha256"] = _sha(
                short_matrix_camera_path
            )
            _write_json(short_matrix_path, short_matrix_manifest)
            with self.assertRaisesRegex(EvaluationError, "16 个有限数值"):
                load_evaluation_manifest(short_matrix_path)

            duplicate_path = _write_manifest(
                root / "duplicate_pose", project_index=2, task_index=1
            )
            duplicate_manifest = json.loads(duplicate_path.read_text(encoding="utf-8"))
            first_camera_path = duplicate_path.parent / "cameras" / "low_left.camera.json"
            second_camera_path = duplicate_path.parent / "cameras" / "low_center.camera.json"
            first_camera = json.loads(first_camera_path.read_text(encoding="utf-8"))
            second_camera = json.loads(second_camera_path.read_text(encoding="utf-8"))
            second_camera["cameraToWorldMatrix"] = first_camera["cameraToWorldMatrix"]
            second_camera["projectionMatrix"] = first_camera["projectionMatrix"]
            _write_json(second_camera_path, second_camera)
            duplicate_manifest["views"][1]["camera"]["sha256"] = _sha(second_camera_path)
            _write_json(duplicate_path, duplicate_manifest)
            with self.assertRaisesRegex(EvaluationError, "重复的相机"):
                load_evaluation_manifest(duplicate_path)

            symlink_path = _write_manifest(
                root / "symlink_escape", project_index=3, task_index=1
            )
            symlink_manifest = json.loads(symlink_path.read_text(encoding="utf-8"))
            external = root / "outside.png"
            external.write_bytes(_png(1, 1, (1, 2, 3)))
            linked_image = symlink_path.parent / "original" / "low_left.png"
            linked_image.unlink()
            linked_image.symlink_to(external)
            symlink_manifest["views"][0]["original"]["sha256"] = _sha(external)
            _write_json(symlink_path, symlink_manifest)
            with self.assertRaisesRegex(EvaluationError, "符号链接"):
                load_evaluation_manifest(symlink_path)

            manifest_link = root / "linked_manifest.json"
            manifest_link.symlink_to(duplicate_path)
            with self.assertRaisesRegex(EvaluationError, "符号链接"):
                load_evaluation_manifest(manifest_link)

            changed_after_plan_path = _write_manifest(
                root / "changed_after_plan", project_index=4, task_index=1
            )
            changed_after_plan = load_evaluation_manifest(changed_after_plan_path)
            planned = plan_gpteval_requests(
                [changed_after_plan],
                evaluator_version="v1",
                model="offline",
                rubric_sha256_value=rubric_sha256("rubric"),
            )
            changed_image = changed_after_plan_path.parent / "inserted" / "high_right.png"
            changed_image.write_bytes(_png(1, 1, (9, 9, 9)))
            late_calls = []
            with self.assertRaisesRegex(EvaluationError, "SHA-256"):
                execute_gpteval_requests(
                    planned,
                    ResponseCache(root / "late-cache"),
                    lambda request: late_calls.append(request.request_key),
                )
            self.assertEqual(late_calls, [])

    def test_supported_offsets_request_keys_and_fake_response_adapter(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_offsets_") as directory:
            root = Path(directory)
            for index, yaw in enumerate((12, 24, 48), start=1):
                path = _write_manifest(
                    root / str(yaw), project_index=index, task_index=1, yaw=yaw
                )
                record = load_evaluation_manifest(path)
                offsets = [view["yawOffsetDegrees"] for view in record.data["views"]]
                self.assertEqual(offsets, [-yaw, 0, yaw, -yaw, 0, yaw])

            rubric_hash = rubric_sha256("rubric")
            sheet_hash = hashlib.sha256(b"sheet").hexdigest()
            base = make_request_key(
                evaluator_version="v1",
                model="model-a",
                rubric_sha256=rubric_hash,
                sheet_sha256=sheet_hash,
                repeat_index=0,
            )
            changed = make_request_key(
                evaluator_version="v1",
                model="model-a",
                rubric_sha256=rubric_hash,
                sheet_sha256=sheet_hash,
                repeat_index=1,
            )
            self.assertNotEqual(base, changed)
            self.assertEqual(
                base,
                canonical_sha256(
                    {
                        "evaluator": "gpteval",
                        "evaluatorVersion": "v1",
                        "model": "model-a",
                        "rubricSha256": rubric_hash,
                        "sheetSha256": sheet_hash,
                        "repeatIndex": 0,
                        "dimensions": list(DEFAULT_DIMENSIONS),
                    }
                ),
            )

            fake = fixed_fake_response(6)
            self.assertEqual(adapt_gpteval_response(fake), fake)
            provider_payload = {
                "candidates": [
                    {"content": {"parts": [{"text": json.dumps(fake)}]}}
                ]
            }
            self.assertEqual(adapt_gpteval_response(provider_payload), fake)
            with self.assertRaises(EvaluationError):
                adapt_gpteval_response({"visual_quality": {"score": 11, "reason": "bad"}})
            with self.assertRaisesRegex(EvaluationError, "严格匹配"):
                adapt_gpteval_response({**fake, "unexpected": {"score": 1, "reason": "bad"}})
            invalid_item = json.loads(json.dumps(fake))
            invalid_item["visual_quality"]["confidence"] = 1
            with self.assertRaisesRegex(EvaluationError, "只能包含"):
                adapt_gpteval_response(invalid_item)
            self.assertEqual(require_supported_evaluator("GPTEval"), "gpteval")
            with self.assertRaisesRegex(EvaluationError, "GPTEval3D_v2"):
                require_supported_evaluator("GPTEval3D_v2")

    def test_dimensions_are_configurable_strict_and_part_of_cache_identity(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_dimensions_") as directory:
            root = Path(directory)
            manifest = load_evaluation_manifest(
                _write_manifest(root / "task", project_index=1, task_index=1)
            )
            common = {
                "evaluator_version": "v1",
                "model": "offline",
                "rubric_sha256_value": rubric_sha256("rubric"),
            }
            default_request = plan_gpteval_requests([manifest], **common)[0]
            three_dimension_request = plan_gpteval_requests(
                [manifest],
                **common,
                dimensions=(
                    "geometric_accuracy",
                    "insertion_rationality",
                    "visual_quality",
                ),
            )[0]

            self.assertEqual(default_request.dimensions, DEFAULT_DIMENSIONS)
            self.assertEqual(three_dimension_request.dimensions, SUPPORTED_DIMENSIONS)
            self.assertNotEqual(
                default_request.request_key,
                three_dimension_request.request_key,
            )
            self.assertEqual(
                normalize_dimensions("geometric_accuracy,visual_quality"),
                DEFAULT_DIMENSIONS,
            )
            self.assertEqual(
                set(fixed_fake_response(7)),
                set(DEFAULT_DIMENSIONS),
            )
            three_scores = fixed_fake_response(7, SUPPORTED_DIMENSIONS)
            self.assertEqual(
                adapt_gpteval_response(three_scores, SUPPORTED_DIMENSIONS),
                three_scores,
            )
            with self.assertRaisesRegex(EvaluationError, "严格匹配"):
                adapt_gpteval_response(three_scores)
            with self.assertRaisesRegex(EvaluationError, "不支持"):
                normalize_dimensions(("visual_quality", "unknown"))

            default_body = build_request_body(default_request, "rubric")
            three_body = build_request_body(three_dimension_request, "rubric")
            default_text = default_body["contents"][0]["parts"][0]["text"]
            three_text = three_body["contents"][0]["parts"][0]["text"]
            self.assertNotIn("insertion_rationality", default_text)
            self.assertIn("insertion_rationality", three_text)
            self.assertEqual(
                default_body["generationConfig"]["responseSchema"]["required"],
                list(DEFAULT_DIMENSIONS),
            )
            self.assertEqual(
                three_body["generationConfig"]["responseSchema"]["required"],
                list(SUPPORTED_DIMENSIONS),
            )

            cache = ResponseCache(root / "three-cache")
            execute_gpteval_requests(
                [three_dimension_request],
                cache,
                lambda request: fixed_fake_response(7, request.dimensions),
            )
            summary = aggregate_gpteval(
                [manifest],
                [three_dimension_request],
                cache,
                expected_tasks_per_method=1,
                expected_scenes_per_method=1,
                expected_tasks_per_scene=1,
            )
            self.assertEqual(summary["dimensions"], list(SUPPORTED_DIMENSIONS))
            default_cache = ResponseCache(root / "default-cache")
            execute_gpteval_requests(
                [default_request],
                default_cache,
                lambda request: fixed_fake_response(7, request.dimensions),
            )
            default_summary = aggregate_gpteval(
                [manifest],
                [default_request],
                default_cache,
                expected_tasks_per_method=1,
                expected_scenes_per_method=1,
                expected_tasks_per_scene=1,
            )
            self.assertNotEqual(
                default_summary["comparisonConfigSha256"],
                summary["comparisonConfigSha256"],
            )
            output = root / "three-summary"
            write_gpteval_summary(output, summary)
            from openpyxl import load_workbook

            workbook = load_workbook(output / "gpteval_summary.xlsx", read_only=True)
            scene_headers = [cell.value for cell in workbook["场景汇总"][1]]
            task_headers = [cell.value for cell in workbook["任务明细"][1]]
            self.assertIn("插入合理性", scene_headers)
            self.assertIn("插入合理性分数", task_headers)
            self.assertIn("插入合理性理由", task_headers)
            task_values = [cell.value for cell in workbook["任务明细"][2]]
            self.assertIn("original/low_left.png", task_values)
            self.assertIn("inserted/high_right.png", task_values)
            self.assertIn("fixed offline test response", task_values)
            self.assertEqual(
                task_values[task_headers.index("任务总分")],
                7,
            )
            workbook.close()

    def test_prompt_cache_concurrency_and_stale_cache_validation(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_cache_") as directory:
            root = Path(directory)
            manifest_path = _write_manifest(root / "task", project_index=1, task_index=1)
            manifest = load_evaluation_manifest(manifest_path)
            rubric_hash = rubric_sha256("rubric")
            request = plan_gpteval_requests(
                [manifest],
                evaluator_version="v1",
                model="offline",
                rubric_sha256_value=rubric_hash,
            )[0]

            changed_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            changed_manifest["taskPrompt"] = "a different requested insertion"
            _write_json(manifest_path, changed_manifest)
            changed_request = plan_gpteval_requests(
                [load_evaluation_manifest(manifest_path)],
                evaluator_version="v1",
                model="offline",
                rubric_sha256_value=rubric_hash,
            )[0]
            self.assertNotEqual(request.request_key, changed_request.request_key)

            request = changed_request
            cache = ResponseCache(root / "cache")
            start = threading.Barrier(3)
            calls = []
            calls_lock = threading.Lock()

            def evaluator(_request):
                with calls_lock:
                    calls.append(_request.request_key)
                time.sleep(0.05)
                return fixed_fake_response(7)

            def run_once():
                start.wait()
                return execute_gpteval_requests([request], cache, evaluator)

            with ThreadPoolExecutor(max_workers=2) as executor:
                futures = [executor.submit(run_once) for _ in range(2)]
                start.wait()
                results = [future.result() for future in futures]
            self.assertEqual(calls, [request.request_key])
            self.assertEqual(sum(item["completed"] for item in results), 1)
            self.assertEqual(sum(item["skipped"] for item in results), 1)

            with self.assertRaisesRegex(EvaluationError, "请求内容不一致"):
                cache.get(replace(request, run_id="stale-run"))
            cache.response_path(request).write_text("{broken", encoding="utf-8")
            with self.assertRaisesRegex(EvaluationError, "缓存损坏"):
                cache.get(request)

            retry_cache = ResponseCache(root / "retry-cache")
            attempts = []

            def fail_once(_request):
                attempts.append(_request.request_key)
                if len(attempts) == 1:
                    raise RuntimeError("fixed test failure")
                return fixed_fake_response(7)

            self.assertEqual(
                execute_gpteval_requests([request], retry_cache, fail_once)["failed"], 1
            )
            self.assertEqual(
                execute_gpteval_requests([request], retry_cache, fail_once)["completed"], 1
            )
            self.assertEqual(
                execute_gpteval_requests([request], retry_cache, fail_once)["skipped"], 1
            )
            self.assertEqual(len(attempts), 2)
            self.assertFalse(retry_cache.error_path(request).exists())

    def test_retry_revalidates_all_inputs_before_another_evaluator_call(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_retry_input_") as directory:
            root = Path(directory)
            manifest_path = _write_manifest(root / "task", project_index=1, task_index=1)
            request = plan_gpteval_requests(
                [load_evaluation_manifest(manifest_path)],
                evaluator_version="v1",
                model="offline",
                rubric_sha256_value=rubric_sha256("rubric"),
            )[0]
            changed_image = manifest_path.parent / "inserted" / "high_right.png"
            calls = []

            class RetryableFailure(RuntimeError):
                retryable = True

            def fail_after_changing_input(current_request):
                calls.append(current_request.request_key)
                changed_image.write_bytes(_png(1, 1, (9, 9, 9)))
                raise RetryableFailure("fixed retryable failure")

            with self.assertRaisesRegex(EvaluationError, "SHA-256"):
                execute_gpteval_requests(
                    [request],
                    ResponseCache(root / "cache"),
                    fail_after_changing_input,
                    retries=1,
                )
            self.assertEqual(calls, [request.request_key])

    def test_request_lock_prevents_duplicate_cross_process_evaluation(self) -> None:
        try:
            process_context = multiprocessing.get_context("fork")
        except ValueError:
            self.skipTest("cross-process flock test requires a fork-capable host")
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_process_") as directory:
            root = Path(directory)
            manifest_path = _write_manifest(root / "task", project_index=1, task_index=1)
            request = plan_gpteval_requests(
                [load_evaluation_manifest(manifest_path)],
                evaluator_version="v1",
                model="offline",
                rubric_sha256_value=rubric_sha256("rubric"),
            )[0]
            start_event = process_context.Event()
            results = process_context.Queue()
            calls = process_context.Queue()
            processes = [
                process_context.Process(
                    target=_cross_process_worker,
                    args=(request, root / "cache", start_event, results, calls),
                )
                for _ in range(2)
            ]
            for process in processes:
                process.start()
            start_event.set()
            progress = [results.get(timeout=10) for _ in processes]
            for process in processes:
                process.join(timeout=10)
                self.assertEqual(process.exitcode, 0)
            self.assertEqual(sum(item["completed"] for item in progress), 1)
            self.assertEqual(sum(item["skipped"] for item in progress), 1)
            self.assertEqual(calls.get(timeout=2), request.request_key)
            with self.assertRaises(queue.Empty):
                calls.get(timeout=0.1)

    def test_partial_denominators_and_equal_weight_aggregation(self) -> None:
        with tempfile.TemporaryDirectory(prefix="insertany3d_eval_weights_") as directory:
            root = Path(directory)
            for task_index in range(1, 6):
                _write_manifest(
                    root / "scene_a" / f"Task_{task_index:03d}",
                    project_index=1,
                    task_index=task_index,
                )
            _write_manifest(root / "scene_b" / "Task_001", project_index=2, task_index=1)
            manifests = discover_evaluation_manifests(root)
            requests = plan_gpteval_requests(
                manifests,
                evaluator_version="v1",
                model="offline",
                rubric_sha256_value=rubric_sha256("rubric"),
            )
            cache = ResponseCache(root / "cache")
            execute_gpteval_requests(
                requests,
                cache,
                lambda request: fixed_fake_response(1 if request.project_id == "Project_001" else 9),
            )
            summary = aggregate_gpteval(
                manifests,
                requests,
                cache,
                expected_tasks_per_method=6,
                expected_scenes_per_method=2,
                expected_tasks_per_scene=5,
            )
            batch_score = summary["methods"]["insertany3d-main"]["batchScore"]
            self.assertEqual(batch_score["sceneMacroAverage"]["visual_quality"]["mean"], 5.0)
            self.assertEqual(
                batch_score["taskMacroAverage"]["visual_quality"]["mean"], 2.333333
            )
            self.assertEqual(summary["status"], "partial")

            single_summary = aggregate_gpteval(
                manifests[:1], requests[:1], cache
            )
            self.assertEqual(single_summary["completion"]["expectedResponses"], 60)
            self.assertEqual(single_summary["completion"]["discoveredResponses"], 1)
            self.assertEqual(single_summary["completion"]["missingResponses"], 59)
            method = single_summary["methods"]["insertany3d-main"]
            self.assertEqual(
                method["batchScore"]["taskMacroAverage"]["visual_quality"]["expectedCount"],
                60,
            )
            self.assertEqual(
                method["batchScore"]["sceneMacroAverage"]["visual_quality"]["expectedCount"],
                12,
            )
            self.assertEqual(
                method["sceneScores"][0]["scores"]["visual_quality"]["expectedCount"],
                5,
            )


def _create_batch(root: Path, *, project_count: int, task_count: int) -> None:
    for project_index in range(1, project_count + 1):
        for task_index in range(1, task_count + 1):
            _write_manifest(
                root / f"Project_{project_index:03d}" / f"Task_{task_index:03d}",
                project_index=project_index,
                task_index=task_index,
            )


def _write_manifest(
    root: Path,
    *,
    project_index: int,
    task_index: int,
    yaw: int = 24,
) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    view_config = {
        "pitchDegrees": [10, 40],
        "yawOffsetDegrees": yaw,
        "layout": "low-high-left-center-right",
    }
    view_config["sha256"] = canonical_sha256(view_config)
    views = []
    image_bytes = _png(1, 1, (project_index % 255, task_index % 255, yaw % 255))
    for view_index, (view_id, pitch, direction) in enumerate(VIEW_LAYOUT):
        original = root / "original" / f"{view_id}.png"
        inserted = root / "inserted" / f"{view_id}.png"
        camera_path = root / "cameras" / f"{view_id}.camera.json"
        original.parent.mkdir(parents=True, exist_ok=True)
        inserted.parent.mkdir(parents=True, exist_ok=True)
        camera_path.parent.mkdir(parents=True, exist_ok=True)
        original.write_bytes(image_bytes)
        inserted.write_bytes(image_bytes)
        _write_json(
            camera_path,
            {
                "viewId": view_id,
                "pitchDegrees": pitch,
                "yawOffsetDegrees": direction * yaw,
                "width": 1,
                "height": 1,
                "cameraToWorld": {
                    "position": {"x": view_index, "y": project_index, "z": task_index},
                    "rotationXyzw": {"x": 0, "y": 0, "z": 0, "w": 1},
                },
                "cameraToWorldMatrixOrder": "row-major",
                "cameraToWorldMatrix": [
                    1, 0, 0, view_index,
                    0, 1, 0, project_index,
                    0, 0, 1, task_index,
                    0, 0, 0, 1,
                ],
                "projectionMatrixOrder": "row-major",
                "projectionMatrix": [
                    1, 0, 0, 0,
                    0, 1, 0, 0,
                    0, 0, 1, view_index + 1,
                    0, 0, 0, 1,
                ],
            },
        )
        views.append(
            {
                "viewId": view_id,
                "pitchDegrees": pitch,
                "yawOffsetDegrees": direction * yaw,
                "original": {
                    "path": f"original/{view_id}.png",
                    "sha256": _sha(original),
                },
                "inserted": {
                    "path": f"inserted/{view_id}.png",
                    "sha256": _sha(inserted),
                },
                "camera": {
                    "path": f"cameras/{view_id}.camera.json",
                    "sha256": _sha(camera_path),
                },
            }
        )
    manifest = {
        "schemaVersion": 1,
        "kind": "insertany3d.evaluation",
        "protocol": "eval6-v1",
        "batchId": "formal-batch",
        "projectId": f"Project_{project_index:03d}",
        "scenePath": f"Assets/Scene_{project_index:03d}.unity",
        "taskId": f"Task_{task_index:03d}",
        "runId": f"run-{project_index:03d}-{task_index:03d}",
        "methodId": "insertany3d-main",
        "taskPrompt": f"insert object {project_index}-{task_index}",
        "viewConfig": view_config,
        "render": {
            "width": 1,
            "height": 1,
            "cameraConvention": "unity-c2w-v1",
        },
        "views": views,
    }
    manifest_path = root / "evaluation_manifest.json"
    _write_json(manifest_path, manifest)
    return manifest_path


def _png(width: int, height: int, color: tuple[int, int, int]) -> bytes:
    def chunk(kind: bytes, payload: bytes) -> bytes:
        return (
            struct.pack(">I", len(payload))
            + kind
            + payload
            + struct.pack(">I", zlib.crc32(kind + payload) & 0xFFFFFFFF)
        )

    row = b"\x00" + bytes(color) * width
    pixels = row * height
    return (
        b"\x89PNG\r\n\x1a\n"
        + chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0))
        + chunk(b"IDAT", zlib.compress(pixels))
        + chunk(b"IEND", b"")
    )


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def _cross_process_worker(request, cache_root, start_event, results, calls) -> None:
    start_event.wait(timeout=5)

    def evaluator(current_request):
        calls.put(current_request.request_key)
        time.sleep(0.2)
        return fixed_fake_response(7)

    results.put(
        execute_gpteval_requests(
            [request], ResponseCache(cache_root), evaluator, retries=0
        )
    )


if __name__ == "__main__":
    unittest.main()
